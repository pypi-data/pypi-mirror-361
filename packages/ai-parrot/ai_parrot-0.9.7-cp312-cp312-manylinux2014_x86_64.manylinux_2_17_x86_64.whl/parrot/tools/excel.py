from typing import Dict, Optional, Type, Any, Literal
import asyncio
import traceback
import uuid
from pathlib import Path
from datetime import datetime
import pandas as pd
from langchain.tools import BaseTool
from pydantic import BaseModel, Field, ConfigDict
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from odf.opendocument import OpenDocumentSpreadsheet
from odf.table import Table, TableRow, TableCell
from odf.text import P
from odf.style import Style, TableCellProperties, TextProperties
from navconfig import BASE_DIR


class ExcelInput(BaseModel):
    """
    Input schema for the ExcelTool. Users can supply:
    • dataframe (required): pandas DataFrame to export
    • output_filename: (Optional) custom filename (including extension) for the generated file
    • sheet_name: (Optional) name of the worksheet (default: "Sheet1")
    • template_file: (Optional) path to Excel/ODS template file to use as base
    • output_format: (Optional) export format - "excel" or "ods" (default: "excel")
    • output_dir: (Optional) directory where the file will be saved
    • header_styles: (Optional) dict of styles to apply to headers
    • data_styles: (Optional) dict of styles to apply to data cells
    """
    model_config = ConfigDict(extra='forbid', arbitrary_types_allowed=True)
    dataframe: Any = Field(..., description="Pandas DataFrame to export to Excel/ODS")
    output_filename: Optional[str] = Field(
        None,
        description="(Optional) Custom filename (including extension) for the generated file"
    )
    sheet_name: Optional[str] = Field(
        "Sheet1",
        description="Name of the worksheet (default: 'Sheet1')"
    )
    template_file: Optional[str] = Field(
        None,
        description="Path to Excel/ODS template file to use as base"
    )
    output_format: Literal["excel", "ods"] = Field(
        "excel",
        description="Export format - 'excel' for .xlsx or 'ods' for OpenDocument"
    )
    output_dir: Optional[str] = Field(
        None,
        description="Directory where the file will be saved"
    )
    header_styles: Optional[Dict[str, Any]] = Field(
        None,
        description="Dict of styles to apply to headers (font, color, etc.)"
    )
    data_styles: Optional[Dict[str, Any]] = Field(
        None,
        description="Dict of styles to apply to data cells"
    )


class ExcelTool(BaseTool):
    """Excel/ODS Generator Tool for exporting pandas DataFrames."""
    name: str = "excel_generator_tool"
    description: str = (
        "Export pandas DataFrames to Excel (.xlsx) or OpenDocument (.ods) files "
        " with custom styling and templates."
        " Supports both Excel and ODS formats, "
        " allows for custom headers, data styles, and templates."
        " Returns the file path of the generated document."
        " Use 'output_format' to specify the desired format ('excel' or 'ods')."
    )
    output_dir: Optional[Path] = None

    args_schema: Type[BaseModel] = ExcelInput

    def __init__(self, output_dir: str = None):
        """Initialize the Excel generator tool."""
        super().__init__()
        self.output_dir = Path(output_dir) if output_dir else BASE_DIR.joinpath(
            "static", "documents", "excel"
        )
        if not self.output_dir.exists():
            self.output_dir.mkdir(parents=True, exist_ok=True)

    async def _arun(self, **kwargs) -> Dict[str, Any]:
        """Async version of the run method."""
        try:
            # Validate input using Pydantic
            input_data = ExcelInput(**kwargs)
            return await self._generate_excel(input_data)
        except Exception as e:
            print(f"❌ Error in ExcelTool._arun: {e}")
            print(traceback.format_exc())
            return {"error": str(e)}

    def _run(self, **kwargs) -> Dict[str, Any]:
        """Synchronous entrypoint."""
        try:
            # Validate input using Pydantic
            input_data = ExcelInput(**kwargs)
        except Exception as e:
            return {"error": f"Invalid input: {e}"}

        try:
            loop = asyncio.get_event_loop()
            if loop.is_running():
                return loop.run_until_complete(self._generate_excel(input_data))
            else:
                return asyncio.run(self._generate_excel(input_data))
        except RuntimeError:
            return asyncio.run(self._generate_excel(input_data))

    async def _generate_excel(self, input_data: ExcelInput) -> Dict[str, Any]:
        """Generate Excel or ODS file from DataFrame."""
        try:
            if input_data.output_format == "excel":
                output_path = self._create_excel_file(input_data)
            else:  # ods
                output_path = self._create_ods_file(input_data)

            return {
                "status": "success",
                "file_path": str(output_path),
                "filename": output_path.name,
                "format": input_data.output_format,
                "sheet_name": input_data.sheet_name,
                "rows": len(input_data.dataframe),
                "columns": len(input_data.dataframe.columns),
                "message": f"{input_data.output_format.upper()} file successfully created at {output_path}"
            }

        except Exception as e:
            return {
                "status": "error",
                "error": str(e),
                "message": f"Failed to create {input_data.output_format.upper()} file: {str(e)}"
            }

    def _create_excel_file(self, input_data: ExcelInput) -> Path:
        """Create Excel file using openpyxl."""
        # Load template or create new workbook
        if input_data.template_file and Path(input_data.template_file).exists():
            wb = load_workbook(input_data.template_file)
            # Clear existing data in the target sheet if it exists
            if input_data.sheet_name in wb.sheetnames:
                ws = wb[input_data.sheet_name]
                ws.delete_rows(1, ws.max_row)
            else:
                ws = wb.create_sheet(input_data.sheet_name)
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = input_data.sheet_name

        # Convert DataFrame to rows
        rows = dataframe_to_rows(input_data.dataframe, index=False, header=True)

        # Add data to worksheet
        for r_idx, row in enumerate(rows, 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)

                # Apply header styles
                if r_idx == 1 and input_data.header_styles:
                    self._apply_excel_cell_style(cell, input_data.header_styles)
                # Apply data styles
                elif r_idx > 1 and input_data.data_styles:
                    self._apply_excel_cell_style(cell, input_data.data_styles)

        # Auto-adjust column widths
        self._adjust_column_widths(ws)

        # Apply default header styling if no custom styles provided
        if not input_data.header_styles:
            self._apply_default_header_styles(ws, len(input_data.dataframe.columns))

        # Save file
        output_path = self._generate_output_path(input_data, "xlsx")
        wb.save(str(output_path))

        return output_path

    def _apply_excel_cell_style(self, cell, styles: Dict[str, Any]):
        """Apply styles to an Excel cell."""
        # Font styling
        if any(key in styles for key in ['font_name', 'font_size', 'bold', 'italic', 'font_color']):
            font_kwargs = {}
            if 'font_name' in styles:
                font_kwargs['name'] = styles['font_name']
            if 'font_size' in styles:
                font_kwargs['size'] = styles['font_size']
            if 'bold' in styles:
                font_kwargs['bold'] = styles['bold']
            if 'italic' in styles:
                font_kwargs['italic'] = styles['italic']
            if 'font_color' in styles:
                font_kwargs['color'] = styles['font_color']

            cell.font = Font(**font_kwargs)

        # Background color
        if 'background_color' in styles:
            cell.fill = PatternFill(
                start_color=styles['background_color'],
                end_color=styles['background_color'],
                fill_type="solid"
            )

        # Alignment
        if any(key in styles for key in ['horizontal', 'vertical', 'wrap_text']):
            align_kwargs = {}
            if 'horizontal' in styles:
                align_kwargs['horizontal'] = styles['horizontal']
            if 'vertical' in styles:
                align_kwargs['vertical'] = styles['vertical']
            if 'wrap_text' in styles:
                align_kwargs['wrap_text'] = styles['wrap_text']

            cell.alignment = Alignment(**align_kwargs)

        # Borders
        if 'border' in styles:
            border_style = styles['border']
            side = Side(style=border_style, color="000000")
            cell.border = Border(left=side, right=side, top=side, bottom=side)

    def _apply_default_header_styles(self, ws, num_columns: int):
        """Apply default styling to header row."""
        header_font = Font(name='Calibri', size=12, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, num_columns + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

    def _adjust_column_widths(self, ws):
        """Auto-adjust column widths based on content."""
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter

            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    def _create_ods_file(self, input_data: ExcelInput) -> Path:
        """Create ODS file using odfpy."""
        # Create new ODS document
        doc = OpenDocumentSpreadsheet()

        # Create styles
        header_style = self._create_ods_header_style(doc, input_data.header_styles)
        data_style = self._create_ods_data_style(doc, input_data.data_styles)

        # Create table (sheet)
        table = Table(name=input_data.sheet_name)

        # Add header row
        header_row = TableRow()
        for col_name in input_data.dataframe.columns:
            cell = TableCell(stylename=header_style)
            cell.addElement(P(text=str(col_name)))
            header_row.addElement(cell)
        table.addElement(header_row)

        # Add data rows
        for _, row in input_data.dataframe.iterrows():
            data_row = TableRow()
            for value in row:
                cell = TableCell(stylename=data_style)
                cell.addElement(P(text=str(value) if pd.notna(value) else ""))
                data_row.addElement(cell)
            table.addElement(data_row)

        doc.spreadsheet.addElement(table)

        # Save file
        output_path = self._generate_output_path(input_data, "ods")
        doc.save(str(output_path))

        return output_path

    def _create_ods_header_style(self, doc, custom_styles: Optional[Dict[str, Any]]) -> str:
        """Create ODS style for headers."""
        style = Style(name="HeaderStyle", family="table-cell")

        # Default header properties
        cell_props = TableCellProperties(
            backgroundcolor="#366092",
            border="1pt solid #000000"
        )
        text_props = TextProperties(
            fontweight="bold",
            color="#FFFFFF",
            fontfamily="Calibri",
            fontsize="12pt"
        )

        # Apply custom styles if provided
        if custom_styles:
            if 'background_color' in custom_styles:
                cell_props.backgroundcolor = custom_styles['background_color']
            if 'font_color' in custom_styles:
                text_props.color = custom_styles['font_color']
            if 'font_name' in custom_styles:
                text_props.fontfamily = custom_styles['font_name']
            if 'font_size' in custom_styles:
                text_props.fontsize = f"{custom_styles['font_size']}pt"
            if 'bold' in custom_styles:
                text_props.fontweight = "bold" if custom_styles['bold'] else "normal"

        style.addElement(cell_props)
        style.addElement(text_props)
        doc.styles.addElement(style)

        return "HeaderStyle"

    def _create_ods_data_style(self, doc, custom_styles: Optional[Dict[str, Any]]) -> str:
        """Create ODS style for data cells."""
        style = Style(name="DataStyle", family="table-cell")

        # Default data properties
        cell_props = TableCellProperties(border="1pt solid #CCCCCC")
        text_props = TextProperties(
            fontfamily="Calibri",
            fontsize="11pt"
        )

        # Apply custom styles if provided
        if custom_styles:
            if 'background_color' in custom_styles:
                cell_props.backgroundcolor = custom_styles['background_color']
            if 'font_color' in custom_styles:
                text_props.color = custom_styles['font_color']
            if 'font_name' in custom_styles:
                text_props.fontfamily = custom_styles['font_name']
            if 'font_size' in custom_styles:
                text_props.fontsize = f"{custom_styles['font_size']}pt"
            if 'bold' in custom_styles:
                text_props.fontweight = "bold" if custom_styles['bold'] else "normal"

        style.addElement(cell_props)
        style.addElement(text_props)
        doc.styles.addElement(style)

        return "DataStyle"

    def _generate_output_path(self, input_data: ExcelInput, extension: str) -> Path:
        """Generate output file path."""
        # Determine output directory
        output_dir = Path(input_data.output_dir) if input_data.output_dir else self.output_dir
        output_dir.mkdir(parents=True, exist_ok=True)

        # Generate filename
        if input_data.output_filename:
            filename = input_data.output_filename
            if not filename.endswith(f'.{extension}'):
                filename = f"{filename.rsplit('.', 1)[0]}.{extension}"
        else:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f"export_{timestamp}_{uuid.uuid4().hex[:8]}.{extension}"

        return output_dir / filename
