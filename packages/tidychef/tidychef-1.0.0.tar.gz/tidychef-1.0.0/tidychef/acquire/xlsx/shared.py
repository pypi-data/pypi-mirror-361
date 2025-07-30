import logging
import re
from typing import Any, Dict, List

from openpyxl.cell.cell import Cell as OpenPyxlCell
from openpyxl.workbook import Workbook

from tidychef.models.source.cell import Cell
from tidychef.models.source.table import Table
from tidychef.selection import Selectable, XlsxSelectable

from ..excel_time import EXCEL_TIME_FORMATS


def xlsx_time_formats():
    """
    Wrapped for test purposes
    """
    return EXCEL_TIME_FORMATS


def sheets_from_workbook(
    source: Any,
    selectable: Selectable,
    workbook: Workbook,
    custom_time_formats: Dict[str, str],
    tables_regex: str,
) -> List[XlsxSelectable]:

    tidychef_selectables = []

    for worksheet_name in workbook.sheetnames:
        time_format_warnings = {}

        if tables_regex is not None:
            if not re.match(tables_regex, worksheet_name):
                continue  # pragma: no cover

        worksheet = workbook[worksheet_name]

        # Unmerge all cells as merged cells do not have the
        # properties we need, i.e .is_date
        for items in sorted(worksheet.merged_cells.ranges):  # pragma: no cover
            worksheet.unmerge_cells(str(items))

        table = Table()
        for y, row in enumerate(worksheet.iter_rows()):

            opycell: OpenPyxlCell
            for x, opycell in enumerate(row):
                if opycell.is_date and opycell.internal_value is not None:
                    strformat_pattern = xlsx_time_formats().get(
                        opycell.number_format, None
                    )
                    if strformat_pattern is None:

                        strformat_pattern = custom_time_formats.get(
                            opycell.number_format, None
                        )

                        if strformat_pattern is None:
                            xy = f"x:{x}, y:{y}"
                            if opycell.number_format not in time_format_warnings:
                                time_format_warnings[opycell.number_format] = []
                            time_format_warnings[opycell.number_format].append(xy)
                            cell_value = opycell.value
                        else:
                            cell_value = opycell.internal_value.strftime(
                                strformat_pattern
                            )
                    else:
                        cell_value = opycell.internal_value.strftime(strformat_pattern)

                elif opycell.value is not None:
                    cell_value = opycell.value
                else:
                    cell_value = ""
                table.add_cell(Cell(x=x, y=y, value=str(cell_value)))

        for bad_fmt, examples in time_format_warnings.items():
            time_issue_cells = (
                ",".join(examples[:5]) if len(examples) > 5 else ",".join(examples)
            )
            logging.warning(
                f"""When processing table "{worksheet_name}" an unknown excel time format "{bad_fmt}" was encountered. Using raw cell value instead. 
For more details on handling excel time formatting see tidychef documentation. Cell(s) in question (max 5 shown): {time_issue_cells}"""
            )

        tidychef_selectables.append(
            selectable(table, source=source, name=worksheet_name)
        )
    return tidychef_selectables
