from unidoc_agent.base_tool import BaseTool
import openpyxl
import json

class ExcelTool(BaseTool):
    def can_handle(self, file_path, mime_type):
        return file_path.endswith(('.xlsx', '.xls'))

    def extract_content(self, file_path):
        wb = openpyxl.load_workbook(file_path)
        structured_data = {}
        for sheet in wb:
            sheet_data = []
            # Get headers from the first row
            headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1)) if cell.value]
            # Process remaining rows
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_dict = {headers[i]: cell for i, cell in enumerate(row) if i < len(headers) and cell is not None}
                if row_dict:  # Only add non-empty rows
                    sheet_data.append(row_dict)
            structured_data[sheet.title] = sheet_data
        return json.dumps(structured_data, indent=2)