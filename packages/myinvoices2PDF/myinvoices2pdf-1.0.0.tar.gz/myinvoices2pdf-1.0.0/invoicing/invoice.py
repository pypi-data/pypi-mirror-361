import os

import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

def generate(invoices_path, pdfs_path, image_path, product_id, product_name,
            amount_purchased,price_per_unit,total_price):
    """
    This function converts invoice Excel files into PDF invoices
    :param invoices_path:
    :param pdfs_path:
    :param image_path:
    :param product_id:
    :param product_name:
    :param amount_purchased:
    :param price_per_unit:
    :param total_price:
    :return:
    """
    filepaths = glob.glob(f"{invoices_path}/*.xlsx")

    for filepath in filepaths:
        filename = Path(filepath).stem

        invoice_nr = filename.split("-")[0]

        invoice_date = filename.split("-")[1].replace(".xlsx", "")

        pdf = FPDF(orientation="P", unit="mm", format="A4")
        pdf.add_page()
        pdf.set_font("Arial", size=12, style="B")

        pdf.cell(w=0, h=10, txt=f"Invoice nr. {invoice_nr}", border=0, align="L", ln=1)
        pdf.cell(w=0, h=10, txt=f"Date: {invoice_date}", border=0, align="L", ln=1)



        df = pd.read_excel(filepath, sheet_name="Sheet 1")
        header_raw = list(df.columns)
        header = [item.replace("_"," ").title() for item in header_raw]

        #Add a header
        pdf.set_font("Arial", size=10, style="B")
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=header[0], border=1)
        pdf.cell(w=70, h=8, txt=header[1], border=1)
        pdf.cell(w=35, h=8, txt=header[2], border=1)
        pdf.cell(w=30, h=8, txt=header[3], border=1)
        pdf.cell(w=30, h=8, txt=header[4], border=1)
        pdf.ln(8)

        total_sum = 0
        for index, row in df.iterrows():

            pdf.set_font("Arial", size=10)
            pdf.set_text_color(80, 80, 80)
            pdf.cell(w=30, h=8, txt=str(row[product_id]), border=1)
            pdf.cell(w=70, h=8, txt=row[product_name], border=1)
            pdf.cell(w=35, h=8, txt=str(row[amount_purchased]), border=1, align="R")
            pdf.cell(w=30, h=8, txt=str(row[price_per_unit]), border=1, align="R")
            pdf.cell(w=30, h=8, txt=str(row[total_price]), border=1, align="R")
            pdf.ln(8)

        total_sum = df[total_price].sum()

        pdf.set_font("Arial", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, border=1)
        pdf.cell(w=70, h=8, border=1)
        pdf.cell(w=35, h=8, border=1)
        pdf.cell(w=30, h=8, border=1)
        pdf.cell(w=30, h=8, txt=str(total_sum), border=1, align="R")
        pdf.ln(8)


        #Add sum
        pdf.set_font("Arial", size=10, style="B")
        pdf.cell(w=0, h=10, txt=f"The total due amount is  {total_sum} Euros.", border=0, align="L", ln=1)

        #Add logo
        pdf.set_font("Arial", size=10, style="B")
        pdf.cell(w=25, h=10, txt="PythonHow", border=0, align="L")
        pdf.image(image_path, w=10)

        if not os.path.exists(pdfs_path):
            os.makedirs(pdfs_path)
        pdf.output(f"{pdfs_path}/{filename.split('/')[-1].split('.')[0]}.pdf")
