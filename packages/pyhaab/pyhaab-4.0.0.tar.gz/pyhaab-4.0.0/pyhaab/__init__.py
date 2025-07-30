"""
PyHaab – Cisco switch inventory collector (Excel-in-place)

Import style
------------
>>> from pyhaab import collect
>>> collect([{"host": "10.0.0.1", "username": "admin", "password": "pwd"}],
...         out_file="Switches.xlsx")

CLI style (installed automatically)
-----------
$ pyhaab 10.0.0.1 admin pwd --out Switches.xlsx
$ pyhaab --csv switches.csv --out Report.xlsx
"""
from pathlib import Path
from typing import List, Dict, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from netmiko import ConnectHandler  # heavy import – only if used

__all__ = ["collect", "collect_from_cli"]

# ╔════════════════════════════════════════════════════════════════╗
# ║                     INTERNAL HELPERS                           ║
# ╚════════════════════════════════════════════════════════════════╝
def _style_switch_sheet(ws):
    widths = (10.4, 11.9, 18.6, 27.9, 32.9, 17, 17)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 16
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name="Bahnschrift", size=12, bold=(r == 1))
            cell.alignment = Alignment(horizontal="center", vertical="center")
    for r in range(2, ws.max_row + 1):
        info, nac = ws.cell(r, 3), ws.cell(r, 7)
        if isinstance(info.value, str):
            v = info.value.lower()
            if v == "up":
                info.fill = PatternFill("solid", fgColor="FFFF00")
            elif "admin" in v:
                info.fill = PatternFill("solid", fgColor="FF5D37")
        if isinstance(nac.value, str):
            v = nac.value.lower()
            if v == "closed":
                nac.fill = PatternFill("solid", fgColor="C6EFCE")
                nac.font = Font(name="Bahnschrift", size=12, color="006100")
            elif v == "monitored":
                nac.fill = PatternFill("solid", fgColor="FFC7CE")
                nac.font = Font(name="Bahnschrift", size=12, color="9C0006")


def _style_summary_sheet(ws):
    widths = (24, 16, 26.5, 12.7, 29, 30, 39, 10, 42, 42)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w
    head_fill = PatternFill("solid", fgColor="305496")
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell.font = Font(name="Bahnschrift", size=12, bold=(r == 1))
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if r == 1:
                cell.fill = head_fill


def _refresh_switch_sheet(
    wb,
    *,
    host: str,
    username: str,
    password: str,
    device_type: str = "cisco_ios",
) -> Dict[str, Dict[str, str]]:
    """
    Return a 1-item dict {hostname: {summary fields…}} and fully
    replace the (hostname) worksheet in *wb*.
    """
    import re  # local import keeps library import-light

    summary: Dict[str, Dict[str, str]] = {}
    try:
        conn = ConnectHandler(
            device_type=device_type, host=host, username=username, password=password
        )
        print(f"[OK]  {host}")

        # hostname
        hostname = next(
            (
                l.split()[1]
                for l in conn.send_command("show run").splitlines()
                if l.startswith("hostname ")
            ),
            host,
        )

        # location
        m = re.search(
            r"snmp-server location\s+(.+)",
            conn.send_command("show run | inc snmp-server location"),
        )
        location = m.group(1).strip() if m else "N/A"

        # show mod
        mod_out = conn.send_command("show mod")
        model, serial, macs, vers = [], [], [], []
        for l in mod_out.splitlines():
            if ("Ports" in l and "Model" in l) or "-----" in l:
                continue
            m = re.match(
                r"^\s*\d+\s+\d+\s+(\S+)\s+(\S+)\s+(\S+)\s+\S+\s+(\S+)", l
            )
            if m:
                model.append(m.group(1))
                serial.append(m.group(2))
                macs.append(m.group(3))
                vers.append(m.group(4))
        if not model:
            model = serial = macs = vers = ["N/A"]
        stack = str(len(serial)) if len(serial) > 1 else "No"
        version = vers[0]

        # uptime
        uptime = next(
            (
                l.split("uptime is", 1)[1].strip()
                for l in conn.send_command("show version").splitlines()
                if "uptime is" in l
            ),
            "N/A",
        )

        # VLAN map + trunk native
        vlan_map, trunk_native, cur_if = {}, {}, None
        for l in conn.send_command("show interface switchport").splitlines():
            if l.startswith("Name:"):
                cur_if = l.split("Name:", 1)[1].strip()
            elif "Access Mode VLAN" in l and cur_if:
                vlan = int(re.search(r"(\d+)", l).group(1))
                vlan_map[cur_if] = "Trunk" if vlan == 1 else str(vlan)
            elif "Trunking Native Mode VLAN" in l and cur_if:
                nat = re.search(r"(\d+)", l).group(1)
                trunk_native[cur_if] = nat
                vlan_map.setdefault(cur_if, "Trunk")

        # interfaces
        int_list = []
        for l in conn.send_command("show interfaces description").splitlines():
            if not l.strip() or l.startswith("Interface"):
                continue
            parts = re.split(r"\s{2,}", l.strip(), 3)
            if len(parts) >= 3 and parts[0].startswith("Gi"):
                iface, status = parts[0], parts[1]
                desc = parts[3] if len(parts) == 4 else ""
                int_list.append(
                    {
                        "Interface": iface,
                        "Status": status,
                        "Device": desc,
                        "VLAN": vlan_map.get(iface, "N/A"),
                    }
                )

        # MAC map
        mac_map = {}
        for tbl in ("dynamic", "static"):
            for l in conn.send_command(
                f"show mac address-table {tbl}"
            ).splitlines():
                if re.match(
                    r"^\s*\d+\s+[0-9a-fA-F]{4}\.[0-9a-fA-F]{4}\.[0-9a-fA-F]{4}",
                    l,
                ):
                    vlan, mac, _typ, port = l.split()[:4]
                    mac_map.setdefault((port, vlan), mac)

        # IP map
        ip_map = {}
        for l in conn.send_command(
            "show device-tracking database"
        ).splitlines():
            if l.startswith(("Network", "Codes", "Preflevel", "---")) or not l.strip():
                continue
            m = re.match(
                r"\s*\S+\s+(\d+\.\d+\.\d+\.\d+)\s+[0-9a-fA-F\.]+\s+(\S+)",
                l,
            )
            if m and not m.group(2).lower().startswith("po"):
                ip_map[m.group(2)] = m.group(1)

        # NAC status
        nac_map = {}
        for item in int_list:
            iface = item["Interface"]
            cfg = conn.send_command(f"show run interface {iface}")
            if "access-session close" in cfg:
                nac_map[iface] = "closed"
            elif "dot1x pae authenticator" in cfg:
                nac_map[iface] = "monitored"
            else:
                nac_map[iface] = "Off"

        # worksheet
        if hostname in wb.sheetnames:
            del wb[hostname]
        ws = wb.create_sheet(hostname)
        ws.append(
            ["Port", "Vlan-ID", "Info", "Device", "MAC", "IP", "NAC"]
        )
        for item in int_list:
            iface, vlan = item["Interface"], item["VLAN"]
            mac = "N/A"
            if vlan.isdigit():
                mac = mac_map.get((iface, vlan), "N/A")
            elif vlan == "Trunk" and iface in trunk_native:
                mac = mac_map.get(
                    (iface, trunk_native[iface]), "N/A"
                )
            ip = ip_map.get(iface, "")
            ws.append(
                [
                    iface,
                    vlan,
                    item["Status"],
                    item["Device"],
                    mac,
                    ip,
                    nac_map[iface],
                ]
            )
        _style_switch_sheet(ws)

        # summary
        summary[hostname] = {
            "IP Address": host,
            "Switch Type": ", ".join(model),
            "Version": version,
            "Model": ", ".join(model),
            "Serial Number": ", ".join(serial),
            "Module MAC": ", ".join(macs),
            "Stack": stack,
            "Location": location,
            "Uptime": uptime,
        }
        conn.disconnect()

    except Exception as exc:  # pragma: no cover
        print(f"[FAIL] {host}: {exc}")
    return summary


def _update_summary_sheet(wb, data):
    hdr = [
        "Switch",
        "IP Address",
        "type",
        "Version",
        "Model",
        "Serial Number",
        "Module MAC",
        "Stack",
        "Location",
        "Uptime",
    ]
    ws = (
        wb["Switches"]
        if "Switches" in wb.sheetnames
        else wb.create_sheet("Switches")
    )
    if ws.max_row == 1:
        ws.append(hdr)

    existing = {ws.cell(r, 1).value: r for r in range(2, ws.max_row + 1)}
    for h, d in data.items():
        row = [
            h,
            d["IP Address"],
            d["Switch Type"],
            d["Version"],
            d["Model"],
            d["Serial Number"],
            d["Module MAC"],
            d["Stack"],
            d["Location"],
            d["Uptime"],
        ]
        if h in existing:
            for c, v in enumerate(row, 1):
                ws.cell(existing[h], c, v)
        else:
            ws.append(row)
    _style_summary_sheet(ws)


def collect(
    switches: List[Dict[str, str]],
    *,
    out_file: str = "00900Switches.xlsx",
) -> Path:
    """
    Programmatic API.  *switches* is a list of::

        {"host": "...", "username": "...", "password": "..."}  dicts

    Returns the saved workbook Path.
    """
    wb = load_workbook(out_file) if Path(out_file).is_file() else Workbook()
    # kill the empty default sheet once
    if wb.active.title == "Sheet" and len(wb.sheetnames) == 1:
        wb.remove(wb.active)

    summary = {}
    for sw in switches:
        summary.update(_refresh_switch_sheet(wb, **sw))

    _update_summary_sheet(wb, summary)
    wb.save(out_file)
    print("UPDATED →", out_file)
    return Path(out_file)


# ╔════════════════════════════════════════════════════════════════╗
# ║                       CLI FRONT-END                           ║
# ╚════════════════════════════════════════════════════════════════╝
def _parse_cli(argv: List[str] | None = None) -> Tuple[List[Dict], str]:
    import argparse, csv, sys

    p = argparse.ArgumentParser(prog="pyhaab")
    p.add_argument(
        "--csv",
        help="CSV with host,username,password columns (header is mandatory)",
    )
    p.add_argument(
        "--out",
        default="00900Switches.xlsx",
        help="Excel file to create/update (default 00900Switches.xlsx)",
    )
    p.add_argument(
        "triples",
        nargs="*",
        help="host user pass triples … (repeat as needed)",
    )
    a = p.parse_args(argv)

    switches: List[Dict[str, str]] = []

    # from CSV
    if a.csv:
        try:
            with open(a.csv, newline="") as f:
                for row in csv.DictReader(f):
                    if not {"host", "username", "password"} <= row.keys():
                        sys.exit("CSV must contain host,username,password")
                    switches.append(
                        {k: row[k].strip() for k in ("host", "username", "password")}
                    )
        except Exception as exc:
            sys.exit(f"[CSV ERROR] {exc}")

    # from positional triples
    if a.triples:
        if len(a.triples) % 3:
            sys.exit("Provide host user pass triples")
        it = iter(a.triples)
        for h, u, pw in zip(it, it, it):
            switches.append({"host": h, "username": u, "password": pw})

    if not switches:
        sys.exit("No switches supplied (use --csv or triples)")

    out = a.out if a.out.lower().endswith(".xlsx") else a.out + ".xlsx"
    return switches, out


def collect_from_cli(argv: List[str] | None = None) -> None:
    switches, out_file = _parse_cli(argv)
    collect(switches, out_file=out_file)


# entry point for -m pyhaab  or the console-script registered in pyproject
def entry() -> None:  # pragma: no cover
    collect_from_cli()


# keep “python -m pyhaab” working too
if __name__ == "__main__":  # pragma: no cover
    entry()
