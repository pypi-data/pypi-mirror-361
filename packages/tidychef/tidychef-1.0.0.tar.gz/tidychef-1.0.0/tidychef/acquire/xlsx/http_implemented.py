"""
Holds the code that defines the local xlsx reader.
"""

import io
from pathlib import Path
from typing import Callable, List, Optional, Union

import openpyxl
import requests
import validators

from tidychef.acquire.base import BaseReader
from tidychef.acquire.main import acquirer
from tidychef.acquire.xlsx.shared import sheets_from_workbook
from tidychef.selection.selectable import Selectable
from tidychef.selection.xlsx.xlsx import XlsxSelectable
from tidychef.utils.http.caching import get_cached_session


def http(
    source: Union[str, Path],
    selectable: Selectable = XlsxSelectable,
    pre_hook: Optional[Callable] = None,
    post_hook: Optional[Callable] = None,
    session: requests.Session = None,
    cache: bool = True,
    tables: str = None,
    **kwargs,
) -> Union[XlsxSelectable, List[XlsxSelectable]]:
    """
    Read data from a Path (or string representing a path)
    present on the same machine where tidychef is running.

    This local xlsx reader uses openpyxl:
    https://openpyxl.readthedocs.io/en/stable/index.html

    Any kwargs passed to this function are propagated to
    the openpyxl.load_workbook() method.

    :param source: A url.
    :param selectable: A class that implements tidychef.selection.selectable.Selectable of an inheritor of. Default is XlsxSelectable
    :param pre_hook: A callable that can take source as an argument
    :param post_hook: A callable that can take the output of XlsxSelectable.parse() as an argument.
    :param session: An optional requests.Session object.
    :param cache: Boolean flag for whether or not to cache get requests.
    :return: A single populated Selectable of type as specified by selectable param.
    """

    assert validators.url(source), f"'{source}' is not a valid http/https url."

    return acquirer(
        source,
        HttpXlsxReader(tables),
        selectable,
        pre_hook=pre_hook,
        post_hook=post_hook,
        session=session,
        cache=cache,
        **kwargs,
    )


class HttpXlsxReader(BaseReader):
    """
    A reader to lead in a source where that source is a locally
    held xlsx file.
    """

    def parse(
        self,
        source: str,
        selectable: Selectable = XlsxSelectable,
        data_only=True,
        session: requests.Session = None,
        cache: bool = True,
        **kwargs,
    ) -> List[XlsxSelectable]:
        """
        Parse the provided source into a list of Selectables. Unless overridden the
        selectable is of type XlsxSelectable.

        Additional **kwargs are propagated to openpyxl.load_workbook()

        :param source: A url.
        :param selectable: The selectable type to be returned.
        :data_only: An openpyxl.load_workbook() option to disable acquisition of non data elements from the tabulated source (macros etc)
        :param session: An optional requests.Session object.
        :param cache: Boolean flag for whether or not to cache get requests.
        :return: A list of type as specified by param selectable.
        """

        if not session:
            if cache:
                session = get_cached_session()
            else:
                session = requests.session()

        response: requests.Response = session.get(source)
        if not response.ok:
            raise requests.exceptions.HTTPError(
                f"""
                Unable to get url: {source}
                {response}
                """
            )

        bio = io.BytesIO()
        bio.write(response.content)
        bio.seek(0)

        custom_time_formats = kwargs.get("custom_time_formats", {})
        kwargs.pop("custom_time_formats", None)

        workbook: openpyxl.Workbook = openpyxl.load_workbook(
            bio, data_only=data_only, **kwargs
        )

        sheets = sheets_from_workbook(
            source, selectable, workbook, custom_time_formats, self.tables
        )

        # In this instance we've filtered the tables at the point of reading, so
        # remove the post load filter.
        self.tables = None

        if len(sheets) == 1:
            return sheets[0]
        return sheets
