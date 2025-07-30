#!/usr/bin/env python3
"""Create test Excel file for testing the converter."""

from datetime import date, timedelta
from decimal import Decimal

import openpyxl


def create_test_excel():
    """Create a test Excel file with multiple sheets."""
    wb = openpyxl.Workbook()

    # Sheet 1: Sales Q1
    ws1 = wb.active
    ws1.title = "Sales_Q1"
    ws1.append(["Order_ID", "Date", "Product", "Quantity", "Price", "Total"])

    base_date = date(2024, 1, 1)
    for i in range(10):
        ws1.append(
            [
                f"ORD{i:04d}",
                base_date + timedelta(days=i),
                f"Product_{i % 3}",
                10 + (i % 5),
                Decimal("99.99") + i,
                f"=D{i+2}*E{i+2}",  # Formula
            ]
        )

    # Sheet 2: Sales Q2 (same structure)
    ws2 = wb.create_sheet("Sales_Q2")
    ws2.append(["Order_ID", "Date", "Product", "Quantity", "Price", "Total"])

    for i in range(10):
        ws2.append(
            [
                f"ORD{i+100:04d}",
                base_date + timedelta(days=i + 90),
                f"Product_{i % 3}",
                15 + (i % 5),
                Decimal("109.99") + i,
                f"=D{i+2}*E{i+2}",  # Formula
            ]
        )

    # Sheet 3: Inventory (different structure)
    ws3 = wb.create_sheet("Inventory")
    ws3.append(["SKU", "Product_Name", "Stock_Level", "Reorder_Point"])

    for i in range(5):
        ws3.append([f"SKU{i:03d}", f"Product_{i}", 100 + i * 10, 25])

    wb.save("test_excel.xlsx")
    print("Created test_excel.xlsx with 3 sheets")
    print("- Sales_Q1: 10 rows with formulas")
    print("- Sales_Q2: 10 rows with formulas (same structure as Q1)")
    print("- Inventory: 5 rows (different structure)")


if __name__ == "__main__":
    create_test_excel()
