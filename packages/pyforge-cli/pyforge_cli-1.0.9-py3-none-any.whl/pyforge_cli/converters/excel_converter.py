"""
Excel to Parquet converter implementation for PyForge.
Supports multi-sheet Excel files with intelligent column signature detection.
"""

import logging
import os
from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, getcontext
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

try:
    import openpyxl

    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

import pyarrow as pa
import pyarrow.parquet as pq
from rich.console import Console
from rich.progress import Progress, TaskID
from rich.table import Table

from .base import BaseConverter

if HAS_OPENPYXL:
    from .excel_table_detector import TableDetector

# Set decimal precision to 27 as per requirements
getcontext().prec = 27

logger = logging.getLogger(__name__)
console = Console()


class ColumnSignature:
    """Represents column signature for comparison across sheets."""

    def __init__(self, columns: List[str]):
        # Normalize column names: strip whitespace and convert to lowercase
        self.columns = [
            col.strip().lower() if col else f"column_{i+1}"
            for i, col in enumerate(columns)
        ]
        self.signature = tuple(self.columns)  # Immutable for hashing
        self.original_columns = columns

    def matches(self, other: "ColumnSignature") -> bool:
        """Check if signatures match (case-insensitive, order-sensitive)."""
        return self.signature == other.signature

    def __hash__(self):
        return hash(self.signature)

    def __eq__(self, other):
        return isinstance(other, ColumnSignature) and self.signature == other.signature

    def __repr__(self):
        return f"ColumnSignature({self.original_columns})"


class ExcelSheetInfo:
    """Information about an Excel sheet."""

    def __init__(
        self,
        name: str,
        headers: List[str],
        row_count: int,
        formula_count: int = 0,
        has_table: bool = True,
        table_confidence: float = 1.0,
        table_issues: List[str] = None,
    ):
        self.name = name
        self.headers = headers
        self.row_count = row_count
        self.formula_count = formula_count
        self.has_table = has_table
        self.table_confidence = table_confidence
        self.table_issues = table_issues or []
        self.column_signature = ColumnSignature(headers) if has_table else None
        self.is_empty = row_count == 0
        self.has_formulas = formula_count > 0


class ExcelAnalysisResult:
    """Results of Excel file analysis."""

    def __init__(self):
        self.file_size_mb: float = 0.0
        self.read_only_mode: bool = False
        self.total_sheets: int = 0
        self.sheets: List[ExcelSheetInfo] = []
        self.valid_sheets: List[ExcelSheetInfo] = []
        self.skipped_sheets: List[ExcelSheetInfo] = []
        self.signature_groups: Dict[ColumnSignature, List[str]] = defaultdict(list)
        self.unique_table_structures: int = 0
        self.total_formulas: int = 0
        self.warnings: List[str] = []


class ExcelConverter(BaseConverter):
    """Excel to Parquet converter with multi-sheet support."""

    def __init__(self):
        super().__init__()
        self.supported_inputs = {".xlsx"}
        self.supported_outputs = {".parquet"}
        if HAS_OPENPYXL:
            self.workbook: Optional[openpyxl.Workbook] = None
            self.table_detector = TableDetector()
        else:
            self.workbook = None
            self.table_detector = None
        self.analysis_result: Optional[ExcelAnalysisResult] = None

    def convert(self, input_path: Path, output_path: Path, **kwargs) -> bool:
        """
        Convert Excel file to Parquet format.

        Args:
            input_path: Path to Excel file
            output_path: Output directory path
            **kwargs: Additional options (combine, separate, etc.)

        Returns:
            bool: True if conversion successful
        """
        try:
            # Check if openpyxl is available
            if not HAS_OPENPYXL:
                console.print(
                    "[red]❌ Excel conversion requires 'openpyxl' package[/red]"
                )
                console.print("[dim]Install with: pip install openpyxl[/dim]")
                return False

            # Validate input
            if not self.validate_input(input_path):
                return False

            # Analyze file
            console.print("🔍 Analyzing Excel file...")
            self.analysis_result = self._analyze_file(str(input_path))

            if not self.analysis_result:
                console.print("[red]❌ Failed to analyze Excel file[/red]")
                return False

            # Display analysis
            self._display_analysis(self.analysis_result)

            # Handle user interaction for conversion strategy
            conversion_strategy = self._get_conversion_strategy(
                self.analysis_result, kwargs
            )

            # Convert based on strategy
            output_files = self._execute_conversion(
                str(input_path), str(output_path), conversion_strategy
            )

            # Display results
            self._display_conversion_results(output_files)

            return True

        except Exception as e:
            logger.error(f"Excel conversion failed: {e}")
            console.print(f"[red]❌ Conversion failed: {e}[/red]")
            return False
        finally:
            if self.workbook:
                self.workbook.close()

    def validate_input(self, input_path: Path) -> bool:
        """Validate input Excel file."""
        if not input_path.exists():
            console.print(f"[red]❌ File not found: {input_path}[/red]")
            return False

        if not input_path.suffix.lower() == ".xlsx":
            console.print("[red]❌ Only .xlsx files are supported[/red]")
            return False

        try:
            # Quick test to see if file can be opened
            test_wb = openpyxl.load_workbook(str(input_path), read_only=True)
            test_wb.close()
        except Exception as e:
            console.print(f"[red]❌ Cannot open Excel file: {e}[/red]")
            return False

        return True

    def _analyze_file(self, input_path: str) -> Optional[ExcelAnalysisResult]:
        """Analyze Excel file structure and content."""
        result = ExcelAnalysisResult()

        # Get file size
        result.file_size_mb = os.path.getsize(input_path) / (1024 * 1024)
        result.read_only_mode = (
            result.file_size_mb > 10
        )  # Use read-only for large files

        console.print(f"📊 File size: {result.file_size_mb:.1f} MB")
        console.print(
            f"📖 Using {'read-only' if result.read_only_mode else 'standard'} mode"
        )

        # Load workbook
        try:
            self.workbook = openpyxl.load_workbook(
                input_path,
                read_only=result.read_only_mode,
                data_only=True,  # Get formula values, not formulas
            )
        except Exception as e:
            console.print(f"[red]❌ Failed to load workbook: {e}[/red]")
            return None

        # Analyze each sheet
        result.total_sheets = len(self.workbook.sheetnames)

        with Progress() as progress:
            task = progress.add_task("Analyzing sheets...", total=result.total_sheets)

            for sheet_name in self.workbook.sheetnames:
                sheet_info = self._analyze_sheet(sheet_name)
                result.sheets.append(sheet_info)

                # Categorize sheets
                if sheet_info.has_table and not sheet_info.is_empty:
                    result.valid_sheets.append(sheet_info)
                    # Group by signature
                    result.signature_groups[sheet_info.column_signature].append(
                        sheet_name
                    )
                else:
                    result.skipped_sheets.append(sheet_info)

                result.total_formulas += sheet_info.formula_count
                progress.update(task, advance=1)

        result.unique_table_structures = len(result.signature_groups)

        # Generate warnings
        if result.total_formulas > 0:
            result.warnings.append(
                f"Detected {result.total_formulas} calculated values. "
                "System will convert all formula results to string."
            )

        # Report skipped sheets
        if result.skipped_sheets:
            skipped_names = [s.name for s in result.skipped_sheets]
            result.warnings.append(
                f"Skipped {len(skipped_names)} sheets with no table structure: {', '.join(skipped_names)}"
            )

            # Add detailed reasons for skipping
            for sheet in result.skipped_sheets:
                if sheet.table_issues:
                    result.warnings.append(
                        f"  - {sheet.name}: {'; '.join(sheet.table_issues)}"
                    )

        return result

    def _analyze_sheet(self, sheet_name: str) -> ExcelSheetInfo:
        """Analyze individual sheet using table detector."""
        sheet = self.workbook[sheet_name]

        # Use table detector to analyze sheet structure
        table_info = self.table_detector.detect_table_structure(sheet, sheet_name)

        # Count formulas if we have a valid table
        formula_count = 0
        if table_info["has_table"] and not self.workbook.read_only:
            start_row = table_info["table_start_row"] + 1  # Skip header
            try:
                if hasattr(sheet, "iter_rows"):
                    for row in sheet.iter_rows(min_row=start_row):
                        for cell in row:
                            if cell.data_type == "f":  # Formula
                                formula_count += 1
                                if formula_count > 100:  # Limit for performance
                                    break
                        if formula_count > 100:
                            break
            except Exception:
                pass  # Continue without formula count if there's an error

        return ExcelSheetInfo(
            name=sheet_name,
            headers=table_info["headers"],
            row_count=table_info["data_rows"],
            formula_count=formula_count,
            has_table=table_info["has_table"],
            table_confidence=table_info["confidence"],
            table_issues=table_info["issues"],
        )

    def _display_analysis(self, result: ExcelAnalysisResult):
        """Display analysis results to user."""
        console.print("✓ File validation passed")
        console.print(f"✓ {result.total_sheets} sheets detected")

        if result.warnings:
            for warning in result.warnings:
                console.print(f"[yellow]⚠️  Warning: {warning}[/yellow]")

        # Create summary table for valid sheets
        if result.valid_sheets:
            table = Table(title="📊 Valid Sheets with Table Structure")
            table.add_column("Sheet", style="cyan")
            table.add_column("Rows", justify="right", style="green")
            table.add_column("Columns", justify="right", style="blue")
            table.add_column("Formulas", justify="right", style="yellow")
            table.add_column("Confidence", justify="right", style="magenta")

            for sheet in result.valid_sheets:
                table.add_row(
                    sheet.name,
                    str(sheet.row_count),
                    str(len(sheet.headers)),
                    str(sheet.formula_count),
                    f"{sheet.table_confidence:.2f}",
                )

            console.print(table)

        # Show skipped sheets if any
        if result.skipped_sheets:
            skipped_table = Table(title="⚠️  Skipped Sheets (No Table Structure)")
            skipped_table.add_column("Sheet", style="red")
            skipped_table.add_column("Reason", style="yellow")

            for sheet in result.skipped_sheets:
                reason = (
                    "; ".join(sheet.table_issues)
                    if sheet.table_issues
                    else "No table structure detected"
                )
                skipped_table.add_row(sheet.name, reason)

            console.print(skipped_table)

        # Show signature analysis
        console.print("✓ Table structure analysis complete")

        if result.unique_table_structures == 0:
            console.print("[red]❌ No valid table structures found[/red]")
        elif result.unique_table_structures == 1:
            console.print("ℹ️  All valid sheets have matching column signatures")
        elif result.unique_table_structures > 1:
            console.print(
                f"⚠️  {result.unique_table_structures} different table structures detected"
            )
            for i, (_signature, sheet_names) in enumerate(
                result.signature_groups.items(), 1
            ):
                console.print(f"   Structure {i}: {sheet_names}")

    def _get_conversion_strategy(
        self, result: ExcelAnalysisResult, kwargs: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Determine conversion strategy based on analysis and user input."""
        strategy = {
            "combine_matching": True,  # Default behavior
            "force_separate": kwargs.get("separate", False),
            "force_combine": kwargs.get("combine", False),
            "auto_proceed": kwargs.get("force", False),
        }

        # If forced options, use them
        if strategy["force_separate"]:
            strategy["combine_matching"] = False
            return strategy

        if strategy["force_combine"]:
            strategy["combine_matching"] = True
            return strategy

        # Interactive decision for matching signatures
        if result.unique_table_structures == 1 and len(result.signature_groups) > 0:
            matching_sheets = list(result.signature_groups.values())[0]
            if len(matching_sheets) > 1:
                console.print("\nMultiple tabs detected with same data model. Options:")
                console.print("[1] Combine into single parquet file (default)")
                console.print("[2] Keep as separate parquet files")

                if not strategy["auto_proceed"]:
                    choice = console.input("Enter choice [1]: ").strip() or "1"
                    strategy["combine_matching"] = choice == "1"

        return strategy

    def _execute_conversion(
        self, input_path: str, output_dir: str, strategy: Dict[str, Any]
    ) -> List[str]:
        """Execute the actual conversion process."""
        output_files = []
        base_name = Path(input_path).stem

        # Ensure output directory exists
        Path(output_dir).mkdir(parents=True, exist_ok=True)

        console.print("\nConverting to Parquet...")

        with Progress() as progress:
            if (
                strategy["combine_matching"]
                and len(self.analysis_result.signature_groups) > 0
            ):
                # Combine sheets with matching signatures
                for (
                    _signature,
                    sheet_names,
                ) in self.analysis_result.signature_groups.items():
                    if len(sheet_names) > 1:
                        output_file = os.path.join(output_dir, f"{base_name}.parquet")
                        task = progress.add_task(
                            f"Combining {len(sheet_names)} sheets...",
                            total=len(sheet_names),
                        )
                        self._convert_combined_sheets(
                            sheet_names, output_file, progress, task
                        )
                        output_files.append(output_file)
                    else:
                        safe_sheet_name = self._sanitize_filename(sheet_names[0])
                        output_file = os.path.join(
                            output_dir, f"{base_name}_{safe_sheet_name}.parquet"
                        )
                        task = progress.add_task(
                            f"Converting {sheet_names[0]}...", total=1
                        )
                        self._convert_single_sheet(sheet_names[0], output_file)
                        progress.update(task, advance=1)
                        output_files.append(output_file)
            else:
                # Convert each valid sheet separately
                task = progress.add_task(
                    "Converting sheets...", total=len(self.analysis_result.valid_sheets)
                )
                for sheet in self.analysis_result.valid_sheets:
                    safe_sheet_name = self._sanitize_filename(sheet.name)
                    output_file = os.path.join(
                        output_dir, f"{base_name}_{safe_sheet_name}.parquet"
                    )
                    self._convert_single_sheet(sheet.name, output_file)
                    output_files.append(output_file)
                    progress.update(task, advance=1)

        return output_files

    def _convert_single_sheet(self, sheet_name: str, output_file: str):
        """Convert a single sheet to Parquet."""
        sheet = self.workbook[sheet_name]
        sheet_info = next(
            s for s in self.analysis_result.valid_sheets if s.name == sheet_name
        )

        # Get table info to know where data starts
        table_info = self.table_detector.detect_table_structure(sheet, sheet_name)
        data_start_row = table_info["table_start_row"] + 1  # Skip header

        # Read all data
        data_rows = []

        if hasattr(sheet, "iter_rows"):
            # Standard mode
            for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                if any(cell is not None for cell in row):  # Skip empty rows
                    # Only take columns that match the header count
                    row_data = row[: len(sheet_info.headers)]
                    data_rows.append(self._convert_row_to_strings(row_data))
        else:
            # Read-only mode
            for idx, row in enumerate(sheet.rows):
                if idx + 1 < data_start_row:  # Skip until data starts
                    continue
                row_values = [cell.value for cell in row[: len(sheet_info.headers)]]
                if any(val is not None for val in row_values):  # Skip empty rows
                    data_rows.append(self._convert_row_to_strings(row_values))

        # Create PyArrow table with all string columns
        self._write_parquet_file(sheet_info.headers, data_rows, output_file)

    def _convert_combined_sheets(
        self, sheet_names: List[str], output_file: str, progress: Progress, task: TaskID
    ):
        """Combine multiple sheets with same structure."""
        all_data = []
        headers = None

        for sheet_name in sheet_names:
            sheet = self.workbook[sheet_name]

            if headers is None:
                sheet_info = next(
                    s for s in self.analysis_result.valid_sheets if s.name == sheet_name
                )
                headers = sheet_info.headers

            # Get table info for this sheet
            table_info = self.table_detector.detect_table_structure(sheet, sheet_name)
            data_start_row = table_info["table_start_row"] + 1  # Skip header

            # Read data from this sheet
            if hasattr(sheet, "iter_rows"):
                for row in sheet.iter_rows(min_row=data_start_row, values_only=True):
                    if any(cell is not None for cell in row):
                        # Only take columns that match the header count
                        row_data = row[: len(headers)]
                        all_data.append(self._convert_row_to_strings(row_data))
            else:
                for idx, row in enumerate(sheet.rows):
                    if idx + 1 < data_start_row:
                        continue
                    row_values = [cell.value for cell in row[: len(headers)]]
                    if any(val is not None for val in row_values):
                        all_data.append(self._convert_row_to_strings(row_values))

            progress.update(task, advance=1)

        self._write_parquet_file(headers, all_data, output_file)

    def _convert_row_to_strings(self, row: Tuple[Any, ...]) -> List[str]:
        """Convert a row of data to strings with proper formatting."""
        converted = []

        for value in row:
            if value is None:
                converted.append("")
            elif isinstance(value, (datetime, date)):
                # Convert to ISO 8601 format
                converted.append(value.isoformat())
            elif isinstance(value, bool):
                # Convert boolean to string
                converted.append(str(value))
            elif isinstance(value, (int, float)):
                # Convert with up to 27 decimal precision
                decimal_value = Decimal(str(value))
                converted.append(str(decimal_value))
            else:
                # Everything else as string
                converted.append(str(value))

        return converted

    def _write_parquet_file(
        self, headers: List[str], data_rows: List[List[str]], output_file: str
    ):
        """Write data to Parquet file."""
        # Create PyArrow schema with all string columns
        schema_fields = [(col, pa.string()) for col in headers]
        schema = pa.schema(schema_fields)

        # Convert to columnar format
        if data_rows:
            columns = list(zip(*data_rows))
        else:
            columns = [[] for _ in headers]

        # Ensure all columns have same length
        max_len = max(len(col) for col in columns) if columns else 0
        columns = [list(col) + [""] * (max_len - len(col)) for col in columns]

        # Create and write table
        table = pa.table(columns, schema=schema)
        pq.write_table(table, output_file, compression="snappy")

    def _display_conversion_results(self, output_files: List[str]):
        """Display conversion results."""
        console.print("\n✅ Conversion Complete!")
        console.print("📁 Output files:")

        total_records = 0
        for file_path in output_files:
            table = pq.read_table(file_path)
            file_records = table.num_rows
            total_records += file_records
            file_size = os.path.getsize(file_path) / 1024  # KB

            console.print(
                f"   - {os.path.basename(file_path)} ({file_records:,} records, {file_size:.1f} KB)"
            )

        if self.analysis_result.total_formulas > 0:
            console.print(
                f"⚠️  {self.analysis_result.total_formulas} formula cells converted to string values"
            )

        console.print(f"📊 Total records: {total_records:,}")

    def _sanitize_filename(self, filename: str) -> str:
        """Create safe filename from sheet name."""
        import re

        # Replace spaces and special characters with underscores
        safe_name = re.sub(r'[<>:"/\\|?*\s]+', "_", filename)
        # Remove multiple consecutive underscores
        safe_name = re.sub(r"_+", "_", safe_name)
        # Remove leading/trailing underscores
        return safe_name.strip("_")

    def get_supported_formats(self) -> List[str]:
        """Return list of supported input formats."""
        return [".xlsx"]

    def get_output_format(self) -> str:
        """Return the output format."""
        return "parquet"
