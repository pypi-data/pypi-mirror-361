#!/usr/bin/env python3
# -*- coding: utf-8 -*-

import os
import datetime as dt
import os.path as osp
from poppy.core.logger import logger
import copy
import re

from poppy.core.configuration import Configuration
from roc.idb.parsers import PALISADEParser
from roc.idb.parsers.idb_elements import Packet
from roc.idb.converters import CDFToFill as Converter
from openpyxl import load_workbook
from roc.film.tools import DESCRIPTOR


__all__ = ["IDBToExcel"]


class IDBToExcel(object):
    """
    Class used to convert the IDB in XML format to the excel format needed by
    the converter from excel to CDF skeleton format by Xavier.
    """

    def __init__(self, args):
        """
        Store arguments from the command line.
        """
        # store command line arguments
        self.args = args

        # create the idb parser # TODO - Add the possibility to use the MIB or PALISADE parser
        self.idb_parser = PALISADEParser(args.idb, args.mapping)

        # read HK dataset from the local descriptor file
        self.dataset = DESCRIPTOR.tasks["make_hk"]["outputs"]

        # get the pipeline descriptor file
        self.descriptor = Configuration.manager["descriptor"]

        # Get the pipeline name
        self.pipeline_id = self.descriptor["pipeline.identifier"]

    def __call__(self):
        """
        Used to start the conversion.
        """
        # get information from XML files
        self.idb_parser.parse()

        # create all packets from the IDB
        self.idb_parser.create_packets_idb()

        # create groups of packets and files from the configuration file
        self.create_groups()

    def create_groups(self):
        """
        From the groups specified in the configuration, read the associated
        packets, retrieve parameters and store them in Excel format files.
        """

        # Check if output directory exists, if it does not, create it
        if not osp.exists(self.args.directory):
            os.mkdir(self.args.directory)

        # loop over hk dataset
        for key, data_descr in self.dataset.items():
            # display
            logger.debug("Treating {0} dataset".format(key))

            # get the workbook of the excel files and the attributes from
            # the HK template
            workbook, attributes = self.base_workbook()

            # update global parameters
            self.update_global(workbook, data_descr)

            # treat the packet defined in the group
            packet_name = data_descr["packet"]

            # get the parameters associated to the packet
            packet = Packet.manager[packet_name]

            # store parameters in the excel file
            self.parameters_to_excel(
                workbook,
                packet.parameters,
                attributes,
            )

            # save the excel file in the output directory
            outfile = osp.join(
                self.args.directory, data_descr["template"].replace(".cdf", ".xlsx")
            )
            logger.info("Saving {0}".format(outfile))
            workbook.save(outfile)

    def get_skeleton_version(self, wb):
        """
        From the workbook given in argument, returns the version of the
        skeleton and validate its format.
        """
        # get global attributes sheet
        sheet = wb["GLOBALattributes"]

        # loop over rows, and if the column name matches, add information
        for row in sheet.iter_rows():
            # name of the attribute
            name = row[0].value

            # if this is the version
            if name == "Skeleton_version":
                # get the version
                version = row[3].value

                if version is None:
                    raise ValueError(
                        "Skeleton version is not set in the HK skeleton " + "template."
                    )

                # remove surrounds
                version = version.replace('"', "")

                # validate the value
                if re.match("^[0-9]{2}$", version) is None:
                    raise ValueError(
                        (
                            "Bad format for the skeleton version in the HK "
                            + "template with value {0}"
                        ).format(version)
                    )

                # else return
                return version

    def update_global(self, wb, group):
        """
        Used to update information in global attributes from what is present
        in the descriptor file.
        """
        # get global attributes sheet
        sheet = wb["GLOBALattributes"]

        # project data from the descriptor
        project = self.descriptor["pipeline.project"]

        # Extract items of the dataset identifier
        # (should look like "SOLO_HK_RPW-<equipment>)
        items = group["identifier"].split("_")

        # loop over rows, and if the column name matches, add information
        for row in sheet.iter_rows():
            # name of the attribute
            name = row[0].value

            # source name
            if name == "Source_name":
                row[3].value = project["source"]

            # logical source
            elif name == "Logical_source":
                row[3].value = "_".join([items[0].lower(), "_HK_", items[2].lower()])

            # descriptor
            elif name == "Descriptor":
                row[3].value = items[2].upper() + ">" + items[2].upper()

            elif name == "MODS":
                row[3].value = dt.datetime.utcnow().isoformat()

            # TEXT field
            elif name == "TEXT":
                row[3].value = (
                    "This file contains RPW {0} housekeeping."
                    + "Parameters are returned in the TM_{0}_HK packets."
                ).format(items[2].split("-")[1])

    def base_workbook(self):
        """
        Create the base workbook of the Excel 2007 format.
        """
        # the name of the file for templates
        template = self.args.hk_template_file

        # create the workbook
        wb = load_workbook(template)

        # get the sheet of variable attributes
        sheet = wb["VARIABLEattributes"]

        # attributes
        attributes = {}

        # loop over rows in the sheet and check if there is parameter templates
        rows = []
        for row in sheet.iter_rows():
            # get the variable name
            variable_name = row[0].value

            # if this is a template
            if variable_name == "PARAMETER_TEMPLATE":
                # add the attribute name and parameters
                attributes[row[1].value] = {
                    "type": row[2].value,
                    "value": row[3].value,
                }
                rows.append(row)

        # delete rows from the template
        for row in rows:
            for cell in row:
                cell.value = None
        # sheet._garbage_collect()

        return wb, attributes

    def parameters_to_excel(self, workbook, parameters, base_attributes):
        """
        Transform parameters into their representation in the Excel file.
        """
        # generate z variables sheet
        self._z_variables_sheet(workbook, parameters)

        # generate variable attributes
        self._variable_attributes_sheet(workbook, parameters, base_attributes)

    def _variable_attributes_sheet(
        self,
        workbook,
        parameters,
        base_attributes,
    ):
        """
        Generate the variable attributes sheet.
        """
        # get the sheet for variable attributes
        sheet = workbook["VARIABLEattributes"]

        # get the converter
        converter = Converter()

        # loop over parameters to add them with information
        for parameter in parameters:
            # make a copy of the structure of attributes
            attributes = copy.copy(base_attributes)

            # type of the parameter
            data_type = parameter.definition.data_type.data_type

            # populate attributes
            attributes["FIELDNAM"]["value"] = parameter.definition.name
            attributes["CATDESC"]["value"] = parameter.definition.description
            attributes["VALIDMIN"]["value"] = (
                parameter.definition.minimum
                if parameter.definition.minimum is not None
                else converter.validmin(data_type)
            )
            attributes["VALIDMIN"]["type"] = data_type
            attributes["VALIDMAX"]["value"] = (
                parameter.definition.maximum
                if parameter.definition.maximum is not None
                else converter.validmax(data_type)
            )
            attributes["VALIDMAX"]["type"] = data_type
            attributes["SCALEMIN"]["type"] = data_type
            attributes["SCALEMAX"]["type"] = data_type
            attributes["SCALEMIN"]["value"] = attributes["VALIDMIN"]["value"]
            attributes["SCALEMAX"]["value"] = attributes["VALIDMAX"]["value"]
            attributes["FILLVAL"]["type"] = data_type
            attributes["FILLVAL"]["value"] = converter(data_type)
            attributes["LABLAXIS"]["value"] = parameter.definition.name
            attributes["SRDB_PARAM_ID"]["value"] = str(parameter)
            if self.idb_parser.is_enumeration(parameter):
                attributes["SRDB_ENUM_ID"]["value"] = self.idb_parser.enumeration_srdb(
                    parameter.definition.data_type,
                    parameter.type,
                )

            # now loop over attributes and add them to the sheet
            for attr_name, values in attributes.items():
                sheet.append(
                    [
                        parameter.definition.name,
                        attr_name,
                        values["type"],
                        values["value"],
                    ]
                )

    def _z_variables_sheet(self, workbook, parameters):
        """
        Responsible for the zVariables sheet.
        """
        # get the sheet of z variables
        sheet = workbook["zVariables"]

        # loop over parameters and add them into the sheet
        for parameter in parameters:
            # add parameter information to the sheet
            sheet.append(
                [
                    parameter.definition.name,
                    parameter.definition.data_type.data_type,
                    1,
                    parameter.definition.data_type.dims(),
                    parameter.definition.data_type.sizes(),
                    parameter.definition.data_type.record_variance(),
                    parameter.definition.data_type.dimension_variances(),
                ]
            )
