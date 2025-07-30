from dotenv import load_dotenv
from langchain_experimental.agents.agent_toolkits import create_pandas_dataframe_agent
from langchain_groq import ChatGroq
from mcp.server.fastmcp import FastMCP
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg')
import os
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment
from pathlib import Path
import pandas as pd
from pydrive.auth import GoogleAuth
from pydrive.drive import GoogleDrive
import re
import seaborn as sns
import traceback



class GoogleDriveHandler:
    def __init__(self,client_config=None, credentials_file=None):
        self.base_path = Path(__file__).resolve().parent
        if client_config is None:
            client_config = next(self.base_path.rglob('client_secrets.json'), None)
        
        if credentials_file is None:
            credentials_file = next(self.base_path.rglob('credentials.json'), None)

        self.client_config =client_config
        self.credentials_file =credentials_file
        self.drive = self.authenticate()

    def authenticate(self):
        gauth = GoogleAuth()
        gauth.LoadClientConfigFile(self.client_config)

        try:
            gauth.LoadCredentialsFile(self.credentials_file)
        except Exception:
            print("No saved credentials found. Proceeding with authentication...")

        if gauth.credentials is None or gauth.access_token_expired:
            gauth.CommandLineAuth()
            gauth.SaveCredentialsFile(self.credentials_file)
        else:
            gauth.Authorize()

        return GoogleDrive(gauth)

    def download_file(self, file_id, file_name, local_folder=None):
        local_folder=self.base_path
        try:
            print("Downloading file...")
            if file_name is None or file_id is None:
                raise ValueError("file_name and file_id are required")

            file = self.drive.CreateFile({'id': file_id})
            file_path = os.path.join(local_folder, file_name)
            os.makedirs(local_folder, exist_ok=True)
            file.GetContentFile(file_path)
            print(f"Downloaded: {file_name}")
            return {"status": "success", "file_path": file_path, "file_name": file_name}
        except Exception as e:
            traceback.print_exc()
            return {"status": "error", "message": str(e)}

    def upload_file(self, file_path, file_name, folder_name):
        try:
            # Check if folder exists
            folder_list = self.drive.ListFile({
                'q': f"title='{folder_name}' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            }).GetList()

            if folder_list:
                folder_id = folder_list[0]['id']
            else:
                folder = self.drive.CreateFile({
                    'title': folder_name,
                    'mimeType': 'application/vnd.google-apps.folder'
                })
                folder.Upload()
                folder_id = folder['id']

            existing_file = self.drive.ListFile({
                'q': f"title='{file_name}' and '{folder_id}' in parents and trashed=false"
            }).GetList()

            if existing_file:
                file = existing_file[0]
            else:
                file = self.drive.CreateFile({
                    'title': file_name,
                    'parents': [{'id': folder_id}]
                })

            file.SetContentFile(file_path)
            file.Upload()

            print(f"Uploaded: {file_name} to folder: {folder_name}")
            return {"status": "success", "file_path": file_path, "file_name": file_name}
        except Exception as e:
            traceback.print_exc()
            return {"status": "error", "message": str(e)}
#format the response
def format_response(response):
    formated=response.replace('.','. ')
    formatted=' '.join(formated.split())
    return formatted
def sanitize_filename(text: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "", text)
def log_to_excel(query, answer, file_path):
    # Ensure the folder exists
    print("excel file is open")
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    new_entry = pd.DataFrame([[query, answer]], columns=(["Query", "Answer"]))
    image_path=os.path.join("plots",f"{query}_plot.png")
    try:
        if os.path.exists(file_path):
            # Get last row using openpyxl
            wb = load_workbook(file_path)
            ws = wb.active
            startrow = ws.max_row + 1

            # Use ExcelWriter to append
            with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
                new_entry.to_excel(writer, index=False, header=False, startrow=startrow-1)
            wb = load_workbook(file_path)
            ws = wb.active

            if os.path.exists(image_path):
                img=ExcelImage(image_path)
                img.width=250
                img.height=180
                img.anchor=f"C{startrow}"
                ws.add_image(img)
             # Wrap text and auto-adjust formatting
            for row in ws.iter_rows(min_row=2, max_col=2,max_row=startrow,min_col=1):  # Columns A & B
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

             # Set column widths (adjust as needed)
            col_widths = {
                'A': 30,  # Query
                'B': 50,  # Answer
                'C': 40   # Plot image space
                }
            for col_letter, width in col_widths.items():
                ws.column_dimensions[col_letter].width = width

            # Optional: increase row height for image visibility
            ws.row_dimensions[startrow].height = 140
            wb.save(file_path)
            return {"status":"success","meassage":str("file written successfully")}
            
        else:
            # Create new Excel file with headers
            new_entry.to_excel(file_path, index=False, engine='openpyxl')

    except Exception as e:
        print(f"Failed to write to Excel: {e}")
        print("Attempting to recreate file...")
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
            new_entry.to_excel(file_path, index=False, engine='openpyxl')
            print("File recreated successfully.")
        except Exception as inner_e:
            print(f"Failed to recreate file: {inner_e}")
#initialize the data and agent
def initialized_data_and_agent():
    global df,agent,save_file_path
    try:
        
        if os.path.exists(save_file_path):
            df=pd.read_excel(save_file_path)
            print("Data loaded successfully")
        else:
            print("data load fail")
            
    except Exception as e:
        print(f"Error loading data: {e}")
        return
        
    print("Initializing agent...")

    #initialize the agent
    try:
        agent=create_pandas_dataframe_agent(
            ChatGroq(model="llama3-70b-8192",temperature=0.4,max_tokens=2000),
            df,
            verbose=False,
            max_iterations=20,          # Increase from default
            max_execution_time=60,
            allow_dangerous_code=True,
            agent_executor_kwargs={"handle_parsing_errors": True, }
            
            )
        print("Agent initialized successfully")
    except Exception as e:
        print(f"Error initializing agent: {e}")
        return


mcp=FastMCP("Data_server")
#load the environment variables
load_dotenv()


drive=GoogleDriveHandler()
#initialize the data and agent
df=None
agent=None
save_file_path=None
save_file_name=None
save_file_id=None

@mcp.tool("analyze_stock_data") 
def analyze_stock_data(query:str,file_id:str="",file_name:str="")->dict:
    global agent,save_file_path,save_file_name,save_file_id
    try:
        if not save_file_id:
            if not file_id or not file_name:
                raise ValueError("file id and Name are required")
            save_file_id=file_id
            save_file_name=file_name
            save_file_path=drive.download_file(save_file_id,save_file_name)['file_path']
              
        
        initialized_data_and_agent()
            
        print("agent initialized")
        if agent is None:
            raise ValueError("Agent is None")
        # Enhance the prompt for better natural-language output
        enhanced_prompt = f"""
        You are a financial data analyst. Explain the result of this analysis query clearly for a non-technical audience. 
        Start with a short summary in plain English, followed by key insights, statistics, or relevant calculations if needed.

        Query: {query}
        """
        response=agent.invoke({"input":enhanced_prompt})
        img_name=f"{query}_plot.png"
        fig = plt.gcf()
        axes = fig.get_axes()
        if fig and axes and any(ax.has_data() for ax in fig.get_axes()):
            try:
                save_path = os.path.join(os.getcwd(),"plots", img_name)
                os.makedirs("plots", exist_ok=True) 
                
                plt.savefig(save_path, bbox_inches='tight')  
                plt.close(fig)
                print(f"Plot saved as {img_name}")

            except Exception as plot_error:
                print("error saving plot: ")
                traceback.print_exc()

        output_text=format_response(response["output"])
        log_to_excel(query,output_text,save_file_path)
        file=drive.upload_file(save_file_path,save_file_name,"Result")
        
        return {"status": "success", "message": str(output_text)}
    except Exception as e:
        return {"status": "error", "message": str(e)}



